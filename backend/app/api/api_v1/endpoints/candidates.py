"""
候选人管理API端点
"""
import uuid
import logging
from datetime import datetime
from typing import List, Optional, Dict, Any

from fastapi import APIRouter, HTTPException, Query, UploadFile, File
from app.services.candidate_service import candidate_service
from app.services.database_service import db_service
from app.models.resume import (
    ResumeInfo, ContactInfo, TaskStatus, UploadTask,
)
from database.models.candidate import CandidateModel

logger = logging.getLogger(__name__)
router = APIRouter()

@router.get("/", response_model=List[Dict[str, Any]], summary="获取所有候选人")
async def get_all_candidates(
    limit: int = Query(100, ge=1, le=1000, description="返回候选人数量限制"),
    offset: int = Query(0, ge=0, description="偏移量")
):
    """
    获取所有候选人列表
    
    - **limit**: 返回候选人数量限制 (1-1000)
    - **offset**: 偏移量，用于分页
    """
    try:
        candidates = await candidate_service.get_all_candidates(limit=limit, offset=offset)
        return [candidate.to_dict() for candidate in candidates]
    except Exception as e:
        raise HTTPException(
            status_code=500,
            detail=f"获取候选人列表失败: {str(e)}"
        )

@router.get("/{candidate_id}", response_model=Dict[str, Any], summary="获取候选人详情")
async def get_candidate(candidate_id: int):
    """
    获取候选人详情
    
    - **candidate_id**: 候选人ID
    """
    try:
        candidate = await candidate_service.get_candidate(candidate_id)
        if not candidate:
            raise HTTPException(
                status_code=404,
                detail="候选人不存在"
            )
        return candidate.to_dict()
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(
            status_code=500,
            detail=f"获取候选人详情失败: {str(e)}"
        )

@router.get("/search/", response_model=List[Dict[str, Any]], summary="搜索候选人")
async def search_candidates(
    name: Optional[str] = Query(None, description="姓名"),
    position: Optional[str] = Query(None, description="职位"),
    status: Optional[str] = Query(None, description="状态"),
    experience_years_min: Optional[int] = Query(None, ge=0, description="最小经验年限"),
    experience_years_max: Optional[int] = Query(None, ge=0, description="最大经验年限"),
    education_level: Optional[str] = Query(None, description="教育水平"),
    skills: Optional[str] = Query(None, description="技能（逗号分隔）"),
    rating_min: Optional[int] = Query(None, ge=1, le=5, description="最小评分"),
    limit: int = Query(100, ge=1, le=1000, description="返回数量限制"),
    offset: int = Query(0, ge=0, description="偏移量")
):
    """
    搜索候选人
    
    - **name**: 姓名关键词
    - **position**: 职位关键词
    - **status**: 候选人状态
    - **experience_years_min**: 最小经验年限
    - **experience_years_max**: 最大经验年限
    - **education_level**: 教育水平
    - **skills**: 技能关键词（逗号分隔）
    - **rating_min**: 最小评分
    - **limit**: 返回数量限制
    - **offset**: 偏移量
    """
    try:
        filters = {}
        if name:
            filters["name"] = name
        if position:
            filters["position"] = position
        if status:
            filters["status"] = status
        if experience_years_min is not None:
            filters["experience_years_min"] = experience_years_min
        if experience_years_max is not None:
            filters["experience_years_max"] = experience_years_max
        if education_level:
            filters["education_level"] = education_level
        if rating_min is not None:
            filters["rating_min"] = rating_min
        
        candidates = await candidate_service.search_candidates(filters, limit, offset)
        
        # 如果指定了技能搜索
        if skills:
            skill_list = [skill.strip() for skill in skills.split(",") if skill.strip()]
            skill_candidates = await candidate_service.get_candidates_by_skills(skill_list, limit, offset)
            # 合并结果并去重
            candidate_ids = set()
            unique_candidates = []
            for candidate in candidates + skill_candidates:
                if candidate.id not in candidate_ids:
                    candidate_ids.add(candidate.id)
                    unique_candidates.append(candidate)
            candidates = unique_candidates[:limit]
        
        return [candidate.to_dict() for candidate in candidates]
    except Exception as e:
        raise HTTPException(
            status_code=500,
            detail=f"搜索候选人失败: {str(e)}"
        )

@router.get("/active/", response_model=List[Dict[str, Any]], summary="获取活跃候选人")
async def get_active_candidates(
    limit: int = Query(100, ge=1, le=1000, description="返回数量限制"),
    offset: int = Query(0, ge=0, description="偏移量")
):
    """
    获取活跃候选人列表
    """
    try:
        candidates = await candidate_service.get_active_candidates(limit, offset)
        return [candidate.to_dict() for candidate in candidates]
    except Exception as e:
        raise HTTPException(
            status_code=500,
            detail=f"获取活跃候选人失败: {str(e)}"
        )

@router.get("/recent/", response_model=List[Dict[str, Any]], summary="获取最近候选人")
async def get_recent_candidates(
    days: int = Query(7, ge=1, le=30, description="最近天数"),
    limit: int = Query(100, ge=1, le=1000, description="返回数量限制")
):
    """
    获取最近添加的候选人
    """
    try:
        candidates = await candidate_service.get_recent_candidates(days, limit)
        return [candidate.to_dict() for candidate in candidates]
    except Exception as e:
        raise HTTPException(
            status_code=500,
            detail=f"获取最近候选人失败: {str(e)}"
        )

@router.get("/top-rated/", response_model=List[Dict[str, Any]], summary="获取高评分候选人")
async def get_top_rated_candidates(
    limit: int = Query(10, ge=1, le=100, description="返回数量限制")
):
    """
    获取评分最高的候选人
    """
    try:
        candidates = await candidate_service.get_top_rated_candidates(limit)
        return [candidate.to_dict() for candidate in candidates]
    except Exception as e:
        raise HTTPException(
            status_code=500,
            detail=f"获取高评分候选人失败: {str(e)}"
        )

@router.put("/{candidate_id}/rate", summary="给候选人评分")
async def rate_candidate(
    candidate_id: int,
    rating: int = Query(..., ge=1, le=5, description="评分 (1-5)")
):
    """
    给候选人评分
    
    - **candidate_id**: 候选人ID
    - **rating**: 评分 (1-5)
    """
    try:
        success = await candidate_service.rate_candidate(candidate_id, rating)
        if not success:
            raise HTTPException(
                status_code=400,
                detail="评分失败，请检查候选人ID和评分值"
            )
        return {"message": "评分成功", "candidate_id": candidate_id, "rating": rating}
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(
            status_code=500,
            detail=f"评分失败: {str(e)}"
        )

@router.put("/{candidate_id}/notes", summary="添加候选人备注")
async def add_candidate_notes(
    candidate_id: int,
    notes: str = Query(..., description="备注内容")
):
    """
    添加候选人备注
    
    - **candidate_id**: 候选人ID
    - **notes**: 备注内容
    """
    try:
        success = await candidate_service.add_candidate_notes(candidate_id, notes)
        if not success:
            raise HTTPException(
                status_code=400,
                detail="添加备注失败，请检查候选人ID"
            )
        return {"message": "备注添加成功", "candidate_id": candidate_id}
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(
            status_code=500,
            detail=f"添加备注失败: {str(e)}"
        )

@router.put("/{candidate_id}/status", summary="更新候选人状态")
async def update_candidate_status(
    candidate_id: int,
    status: str = Query(..., description="状态 (active, inactive, hired, rejected)")
):
    """
    更新候选人状态
    
    - **candidate_id**: 候选人ID
    - **status**: 状态
    """
    try:
        success = await candidate_service.update_candidate_status(candidate_id, status)
        if not success:
            raise HTTPException(
                status_code=400,
                detail="更新状态失败，请检查候选人ID和状态值"
            )
        return {"message": "状态更新成功", "candidate_id": candidate_id, "status": status}
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(
            status_code=500,
            detail=f"更新状态失败: {str(e)}"
        )

@router.get("/statistics/", response_model=Dict[str, Any], summary="获取候选人统计信息")
async def get_candidate_statistics():
    """
    获取候选人统计信息
    """
    try:
        stats = await candidate_service.get_candidate_statistics()
        return stats
    except Exception as e:
        raise HTTPException(
            status_code=500,
            detail=f"获取统计信息失败: {str(e)}"
        )

@router.get("/skills/statistics/", response_model=Dict[str, int], summary="获取技能统计信息")
async def get_skills_statistics():
    """
    获取技能统计信息
    """
    try:
        stats = await candidate_service.get_skills_statistics()
        return stats
    except Exception as e:
        raise HTTPException(
            status_code=500,
            detail=f"获取技能统计失败: {str(e)}"
        )

@router.put("/{candidate_id}", summary="更新候选人信息")
async def update_candidate(candidate_id: int, updates: Dict[str, Any]):
    """
    更新候选人基本信息（姓名、联系方式、教育、技能等）
    """
    try:
        candidate = await candidate_service.get_candidate(candidate_id)
        if not candidate:
            raise HTTPException(status_code=404, detail="候选人不存在")
        
        allowed_fields = [
            "name", "phone", "email", "address", "position",
            "education_level", "school", "major", "skills",
            "summary", "notes", "status"
        ]
        for key, value in updates.items():
            if key in allowed_fields and hasattr(candidate, key):
                if key == "skills" and isinstance(value, list):
                    import json
                    value = json.dumps(value, ensure_ascii=False)
                setattr(candidate, key, value)
        
        from datetime import datetime
        candidate.updated_at = datetime.now()
        success = db_service.update_candidate(candidate)
        if not success:
            raise HTTPException(status_code=500, detail="更新失败")
        
        return candidate.to_dict()
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"更新候选人失败: {str(e)}")

@router.delete("/{candidate_id}", summary="删除候选人")
async def delete_candidate(candidate_id: int):
    """
    删除候选人
    
    - **candidate_id**: 候选人ID
    """
    try:
        success = await candidate_service.delete_candidate(candidate_id)
        if not success:
            raise HTTPException(
                status_code=404,
                detail="候选人不存在"
            )
        return {"message": "候选人删除成功", "candidate_id": candidate_id}
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(
            status_code=500,
            detail=f"删除候选人失败: {str(e)}"
        )


@router.post("/import-excel", summary="从 Excel 批量导入候选人")
async def import_candidates_from_excel(file: UploadFile = File(...)):
    """
    从 Excel 文件批量导入候选人（轻量导入，不走简历解析）。

    Excel 格式要求：第一行为表头，包含「姓名」「电话」列，其余列内容合并写入备注。
    """
    if not file.filename.endswith((".xlsx", ".xls")):
        raise HTTPException(status_code=400, detail="仅支持 .xlsx / .xls 格式")

    try:
        import openpyxl
        from io import BytesIO

        content = await file.read()
        wb = openpyxl.load_workbook(BytesIO(content), read_only=True)
        ws = wb.active

        headers = [str(cell.value).strip() if cell.value else "" for cell in next(ws.iter_rows(min_row=1, max_row=1))]
        name_col = _find_col(headers, ["姓名", "名字", "name"])
        phone_col = _find_col(headers, ["电话", "手机", "phone", "tel"])

        if name_col is None:
            raise HTTPException(status_code=400, detail="未找到「姓名」列，请检查表头")

        note_cols = [
            i for i in range(len(headers))
            if i != name_col and i != phone_col and headers[i]
        ]

        success_count = 0
        skip_count = 0
        errors: List[str] = []

        for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            row = list(row)
            name = str(row[name_col]).strip() if name_col < len(row) and row[name_col] else ""
            if not name:
                skip_count += 1
                continue

            phone = str(row[phone_col]).strip() if phone_col is not None and phone_col < len(row) and row[phone_col] else None
            if phone:
                phone = phone.replace(".0", "")

            note_parts = []
            for ci in note_cols:
                if ci < len(row) and row[ci]:
                    val = str(row[ci]).strip()
                    if val:
                        note_parts.append(f"{headers[ci]}: {val}")
            notes_text = "\n".join(note_parts) if note_parts else None

            try:
                task_id = str(uuid.uuid4())
                task = UploadTask(
                    id=task_id,
                    filename=f"excel_import_{file.filename}",
                    file_path="",
                    file_size=0,
                    file_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    status=TaskStatus.UPLOADED,
                )
                db_service.create_task(task)

                resume_info = ResumeInfo(
                    name=name,
                    contact=ContactInfo(phone=phone) if phone else None,
                    other=notes_text,
                )
                db_service.update_task_status(
                    task_id, TaskStatus.COMPLETED, progress=100, result=resume_info
                )
                success_count += 1
            except Exception as e:
                logger.warning("导入第 %d 行失败: %s", row_idx, e)
                errors.append(f"第{row_idx}行({name}): {str(e)}")

        wb.close()

        return {
            "message": f"导入完成：成功 {success_count} 条，跳过 {skip_count} 条",
            "success": success_count,
            "skipped": skip_count,
            "errors": errors[:20],
        }
    except HTTPException:
        raise
    except Exception as e:
        logger.error("Excel 导入失败: %s", e, exc_info=True)
        raise HTTPException(status_code=500, detail=f"导入失败: {str(e)}")


def _find_col(headers: List[str], candidates: List[str]) -> Optional[int]:
    """在表头中按关键词模糊匹配列索引"""
    for i, h in enumerate(headers):
        hl = h.lower()
        for kw in candidates:
            if kw.lower() in hl:
                return i
    return None
